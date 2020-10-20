import openpyxl

def write_to_out_sheet(question, answer, pid, out_sheet):
    pass


excel = openpyxl.load_workbook("C://Users//Ray//Desktop//化工安全技术朱旭--4993(测).xlsx")

data_sheet = excel["企业级问答"]

#print(data_sheet['A1'].value)

out_sheet = excel.create_sheet("neo4j0", 1)
answer = ''
pid = 0

for i in range(20, data_sheet.max_row):
    ##这个写具体筛选逻辑例如:
    question = data_sheet['D{}'.format(i)].value

    if data_sheet['E{}'.format(i)].value != '':
        answer = data_sheet['E{}'.format(i)].value
    write_to_out_sheet(question, answer, pid, out_sheet)
    pid += 1